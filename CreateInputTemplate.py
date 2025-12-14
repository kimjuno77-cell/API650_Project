
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_template(filename="API650_Input_Template.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input"
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Dark Blue
    
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # Light Yellow
    calc_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # Grey
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def set_cell(row, col, value, is_header=False, is_input=False, comment=None):
        c = ws.cell(row=row, column=col, value=value)
        c.border = thin_border
        if is_header:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
        elif is_input:
            c.fill = input_fill
            c.alignment = Alignment(horizontal='center')
        
    # --- 1. General Geometry (Col F / 6) ---
    # Labels
    set_cell(14, 5, "Nominal Diameter (D, mm):", is_header=True) # E14
    set_cell(15, 5, "Tank Height (H, mm):", is_header=True)
    set_cell(16, 5, "Design Liquid Level (Hd, mm):", is_header=True)
    set_cell(17, 5, "Test Liquid Level (Ht, mm):", is_header=True)
    set_cell(24, 5, "Specific Gravity (G):", is_header=True)
    set_cell(32, 5, "Design Pressure (mmH2O):", is_header=True)
    set_cell(34, 5, "Test Pressure (mmH2O):", is_header=True)
    
    # Inputs
    set_cell(14, 6, 20000, is_input=True) # D
    set_cell(15, 6, 12000, is_input=True) # H
    set_cell(16, 6, 12000, is_input=True) # Hd
    set_cell(17, 6, 10000, is_input=True) # Ht
    set_cell(24, 6, 1.0, is_input=True)   # G
    set_cell(32, 6, 250, is_input=True)   # P_design
    set_cell(34, 6, 300, is_input=True)   # P_test

    # --- 2. Corrosion Allowance (Col D / 4) ---
    set_cell(39, 3, "Shell C.A. (mm):", is_header=True) # C39
    set_cell(39, 4, 1.5, is_input=True)
    set_cell(45, 3, "Bottom C.A. (mm):", is_header=True) # C45
    set_cell(45, 4, 3.0, is_input=True)

    # --- 3. Stresses (Col C / 3) ---
    set_cell(56, 2, "Design Stress Sd (MPa):", is_header=True) # B56
    set_cell(56, 3, 160, is_input=True)
    set_cell(57, 2, "Test Stress St (MPa):", is_header=True)
    set_cell(57, 3, 171, is_input=True)

    # --- 4. Wind Load (Col O / 15) ---
    # Labels
    set_cell(12, 14, "Basic Wind Speed (km/h):", is_header=True) # N12
    set_cell(19, 14, "3-sec Gust Speed (m/s):", is_header=True) # N19
    set_cell(21, 14, "Kzt:", is_header=True)
    set_cell(22, 14, "Kd:", is_header=True)
    set_cell(23, 14, "Gust Factor G:", is_header=True)
    set_cell(24, 14, "Force Coeff Cf:", is_header=True)

    # Inputs
    set_cell(12, 15, 0, is_input=True)    # Basic
    set_cell(19, 15, 45.0, is_input=True) # 3-sec
    set_cell(21, 15, 1.0, is_input=True)  # Kzt
    set_cell(22, 15, 0.95, is_input=True) # Kd
    set_cell(23, 15, 0.85, is_input=True) # G
    set_cell(24, 15, 0.5, is_input=True)  # Cf

    # --- 5. Seismic Load (Various Cols) ---
    # Site Class (Row 31, Col S / 19)
    set_cell(31, 18, "Site Class:", is_header=True) # R31
    set_cell(31, 19, "D", is_input=True)

    # Importance Factor I (Row 32, Col U / 21)
    set_cell(32, 20, "Importance Factor I:", is_header=True) # T32
    set_cell(32, 21, 1.0, is_input=True)
    
    # Zone Z (Row 29, Col U / 21)
    set_cell(29, 20, "Zone Coefficient Z:", is_header=True)
    set_cell(29, 21, 0.0, is_input=True)

    # S1 (Row 42, Col Q / 17)
    set_cell(42, 16, "S1:", is_header=True) # P42
    set_cell(42, 17, 0.28, is_input=True)
    
    # S0 (Row 43, Col Q / 17 if used... code read Row 42 for S0, let's check) 
    # Reader: S0 = get_val(42, 16) -> Row 43, Col Q
    set_cell(43, 16, "S0 (Ss?):", is_header=True)
    set_cell(43, 17, 0.5, is_input=True)

    # SDS (Row 46, Col S / 19)
    set_cell(46, 18, "SDS:", is_header=True)
    set_cell(46, 19, 0.48, is_input=True)

    # --- 6. Roof Data ---
    # Searches for 'Roof Plate' in Col B (2). Let's put it at Row 50.
    set_cell(50, 2, "Roof Plate", is_header=True)
    
    set_cell(49, 4, "Roof C.A. (mm)", is_header=True)
    set_cell(50, 4, 0.0, is_input=True)
    
    set_cell(49, 6, "Roof Material", is_header=True) # Col F
    set_cell(50, 6, "A 36", is_input=True)
    
    set_cell(49, 15, "Roof Thk (mm)", is_header=True) # Col O (15) for Thickness? Reader checks 37, 16, 14. 14 is N(14)? No 14 is N. 15 is O. 
    # Reader: get_val(row, 14) -> Col O (15).
    set_cell(50, 15, 6.0, is_input=True)

    # --- 7. Shell Courses Table ---
    # Start Row 161
    set_cell(160, 12, "Course Name (Include 'Shell')", is_header=True) # L160
    set_cell(160, 14, "Material", is_header=True) # N160
    set_cell(160, 15, "Thickness (mm)", is_header=True) # O160
    set_cell(160, 17, "Width (mm)", is_header=True) # Q160
    
    courses = [
        ("1st Shell Ring", "A 573 70", 22, 2438),
        ("2nd Shell Ring", "A 573 70", 18, 2438),
        ("3rd Shell Ring", "A 283 C", 12, 2438),
        ("4th Shell Ring", "A 283 C", 10, 2438),
    ]
    
    start_row = 161
    for i, c in enumerate(courses):
        r = start_row + i
        set_cell(r, 12, c[0], is_input=True)
        set_cell(r, 14, c[1], is_input=True)
        set_cell(r, 15, c[2], is_input=True)
        set_cell(r, 17, c[3], is_input=True)
        
    wb.save(filename)
    print(f"Template created: {filename}")

if __name__ == "__main__":
    create_template()
